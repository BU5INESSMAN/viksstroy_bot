"""SMR Excel report generator.

Produces a clean multi-sheet .xlsx for a single application with:
    - "Объекты": every application/object in a merged SMR
    - "Часы" (hours): object | application | brigade | FIO | hours | filled by
    - "Работы" (plan works): object | application | name | unit | volume | filled by
    - "Доп. работы" (extras): same columns (sheet omitted when empty)

Catalog prices stay role-gated. Participant pay is an independent optional
column and never mixes with the catalog work-rate salary. The "Заполнил"
column uses `application_kp.filled_by_user_id` to resolve
the submitter's FIO + role.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from application_numbers import display_application_number
from smr_calculations import calculate_smr_totals


_HEADER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ROLE_RU = {
    'foreman': 'Прораб',
    'brigadier': 'Бригадир',
    'worker': 'Рабочий',
    'driver': 'Водитель',
    'moderator': 'Модератор',
    'boss': 'Руководитель',
    'superadmin': 'Админ',
}


def _sanitize_filename(name: str) -> str:
    """Windows-safe filename stem — Russian letters preserved."""
    if not name:
        return 'report'
    cleaned = re.sub(r'[\\/:*?"<>|]+', '_', name)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip(' ._')
    return cleaned[:80].rstrip(' .') or 'report'


def _format_work_date(value: str | None) -> str:
    """Format an ISO work date as the Russian DD.MM.YYYY filename form."""
    raw = str(value or '').strip()
    if raw:
        try:
            return datetime.strptime(raw[:10], '%Y-%m-%d').strftime('%d.%m.%Y')
        except ValueError:
            # Keep an already formatted Russian date readable.
            if re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', raw):
                return raw
    return datetime.now().strftime('%d.%m.%Y')


def _build_report_filename(
    *, object_name: str | None, app_id: int,
    public_number: str | None, date_target: str | None,
    team_name: str | None = None,
) -> str:
    obj_name = _sanitize_filename(object_name or f"Объект {app_id}")
    app_number = _sanitize_filename(display_application_number(app_id, public_number))
    work_date = _format_work_date(date_target)
    team_suffix = f" - {_sanitize_filename(team_name)}" if team_name else ""
    return f"{obj_name} - {app_number} - {work_date}{team_suffix}.xlsx"


def _author_label(fio: str | None, role: str | None) -> str:
    if not fio:
        return ''
    role_ru = _ROLE_RU.get((role or '').strip(), '')
    return f"{fio} ({role_ru})" if role_ru else fio


def _row_context(row: dict, app_meta: dict) -> tuple[str, str]:
    return (
        str(row.get('object_name') or app_meta.get('object_name') or app_meta.get('object_address') or ''),
        str(row.get('application_label') or app_meta.get('public_number') or f"№{app_meta.get('id', '')}"),
    )


def _autosize(ws, min_w: int = 12, max_w: int = 60):
    for col_idx, column in enumerate(ws.columns, start=1):
        longest = 0
        for cell in column:
            if cell.value is None:
                continue
            for line in str(cell.value).split('\n'):
                longest = max(longest, len(line))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(longest + 2, max_w))


def _write_header(ws, headers: list[str]):
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=text)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'


async def generate_smr_excel_bytes(
    db,
    app_id: int,
    *,
    include_financial: bool = False,
    include_participant_salary: bool = False,
    team_id: int | None = None,
    team_name: str | None = None,
    include_unassigned_for_team: bool = False,
) -> tuple[bytes, str]:
    """Generate the SMR report .xlsx in memory.
    Returns (file_bytes, suggested_filename).
    """
    # ── Application + object for filename + header context ──
    async with db.conn.execute(
        """
        SELECT a.id, a.public_number, a.date_target, a.foreman_name, a.object_id,
               o.name AS object_name, o.address AS object_address
        FROM applications a
        LEFT JOIN objects o ON o.id = a.object_id
        WHERE a.id = ?
        """,
        (app_id,),
    ) as cur:
        row = await cur.fetchone()
    app_meta = dict(row) if row else {}

    from smr_data import get_smr_read_model

    report = await get_smr_read_model(db, app_id)
    if team_id is not None:
        target_team_id = int(team_id)

        def belongs_to_team(item: dict) -> bool:
            try:
                item_team_id = int(item.get('team_id') or 0)
            except (TypeError, ValueError):
                item_team_id = 0
            return item_team_id == target_team_id or (
                include_unassigned_for_team and item_team_id == 0
            )

        def normalize_legacy_team(item: dict) -> dict:
            if not include_unassigned_for_team or item.get('team_id'):
                return item
            return {
                **item,
                'team_id': target_team_id,
                'team_name': team_name or f'Бригада {target_team_id}',
            }

        report = {**report}
        report['hours'] = [
            normalize_legacy_team(row)
            for row in report.get('hours', []) if belongs_to_team(row)
        ]
        report['plan_works'] = [
            normalize_legacy_team(row)
            for row in report.get('plan_works', []) if belongs_to_team(row)
        ]
        report['extra_works'] = [
            normalize_legacy_team(row)
            for row in report.get('extra_works', []) if belongs_to_team(row)
        ]

        source_application_ids = set()
        for key in ('hours', 'plan_works', 'extra_works'):
            for row in report.get(key, []):
                try:
                    source_id = int(row.get('source_application_id') or row.get('application_id') or 0)
                except (TypeError, ValueError):
                    continue
                if source_id > 0:
                    source_application_ids.add(source_id)
        if source_application_ids:
            filtered_contexts = []
            for context in report.get('applications', []):
                try:
                    context_id = int(context.get('id') or 0)
                except (TypeError, ValueError):
                    continue
                if context_id in source_application_ids:
                    filtered_contexts.append(context)
            report['applications'] = filtered_contexts
        report['totals'] = calculate_smr_totals(
            report['plan_works'], report['extra_works'], report['hours']
        )
    wb = Workbook()

    applications = report.get('applications') or []
    include_object_context = len(applications) > 1
    if include_object_context:
        ws_objects = wb.active
        ws_objects.title = "Объекты"
        _write_header(ws_objects, ["Заявка", "Объект", "Адрес", "Дата работ"])
        for row_no, context in enumerate(applications, start=2):
            values = (
                context.get('application_label') or f"№{context.get('id')}",
                context.get('object_name') or '',
                context.get('object_address_clean') or context.get('object_address') or '',
                _format_work_date(context.get('date_target')),
            )
            for col, value in enumerate(values, start=1):
                ws_objects.cell(row=row_no, column=col, value=value).alignment = _LEFT
        _autosize(ws_objects)
        ws_hours = wb.create_sheet("Часы")
    else:
        ws_hours = wb.active
        ws_hours.title = "Часы"

    # ─────────────────────── Sheet 1: Часы ───────────────────────
    hours_headers = []
    if include_object_context:
        hours_headers.extend(["Объект", "Заявка"])
    hours_headers.extend(["Бригада", "ФИО", "Специальность", "Часы"])
    if include_participant_salary:
        hours_headers.append("ЗП участника")
    hours_headers.append("Заполнил")
    _write_header(ws_hours, hours_headers)

    # v2.10: include addendum hours (доп.отчёт) so they count in the report,
    # matching the works/extras sheets which read the tables directly.
    hours_rows = sorted(
        report.get('hours', []),
        key=lambda row: (str(row.get('object_name') or ''), str(row.get('team_name') or ''), str(row.get('fio') or '')),
    )
    r = 2
    for h in hours_rows:
        if (
            float(h.get('hours') or 0) <= 0
            and float(h.get('participant_salary') or 0) <= 0
        ):
            continue
        col = 1
        if include_object_context:
            object_name, application_label = _row_context(h, app_meta)
            ws_hours.cell(row=r, column=col, value=object_name).alignment = _LEFT
            ws_hours.cell(row=r, column=col + 1, value=application_label).alignment = _LEFT
            col += 2
        ws_hours.cell(row=r, column=col, value=h.get('team_name') or '').alignment = _LEFT
        ws_hours.cell(row=r, column=col + 1, value=h.get('fio') or '').alignment = _LEFT
        ws_hours.cell(row=r, column=col + 2, value=h.get('specialty') or '').alignment = _LEFT
        ws_hours.cell(row=r, column=col + 3, value=float(h.get('hours') or 0)).alignment = _CENTER
        col += 4
        if include_participant_salary:
            ws_hours.cell(
                row=r, column=col,
                value=round(float(h.get('participant_salary') or 0), 2),
            ).alignment = _CENTER
            col += 1
        ws_hours.cell(
            row=r, column=col,
            value=_author_label(h.get('filled_by_fio'), h.get('filled_by_role')),
        ).alignment = _LEFT
        r += 1
    if r == 2:
        ws_hours.cell(row=2, column=1, value='Часы не заполнены').alignment = _LEFT
    _autosize(ws_hours)

    # ─────────────────────── Sheet 2: Работы ───────────────────────
    # v2.4.3 per-brigade: when any row carries team_id, prepend a "Бригада"
    # column so the distribution across teams is visible in the report.
    works = sorted(
        report.get('plan_works', []),
        key=lambda row: (str(row.get('object_name') or ''), str(row.get('category') or ''), str(row.get('name') or '')),
    )

    has_teams = any(w.get('team_id') for w in works)
    ws_works = wb.create_sheet("Работы")
    headers = ["Объект", "Заявка"] if include_object_context else []
    if has_teams:
        headers.extend(["Бригада", "Наименование", "Ед.изм", "Объём"])
    else:
        headers.extend(["Наименование", "Ед.изм", "Объём"])
    if include_financial:
        headers.extend(["ЗП за ед.", "Сумма ЗП", "Цена за ед.", "Сумма цены"])
    headers.append("Заполнил")
    _write_header(ws_works, headers)

    r = 2
    for w in works:
        col = 1
        if include_object_context:
            object_name, application_label = _row_context(w, app_meta)
            ws_works.cell(row=r, column=col, value=object_name).alignment = _LEFT
            ws_works.cell(row=r, column=col + 1, value=application_label).alignment = _LEFT
            col += 2
        if has_teams:
            ws_works.cell(row=r, column=col, value=w.get('team_name') or '—').alignment = _LEFT
            col += 1
        ws_works.cell(row=r, column=col, value=w.get('name') or '').alignment = _LEFT
        col += 1
        ws_works.cell(row=r, column=col, value=w.get('unit') or '').alignment = _CENTER
        col += 1
        ws_works.cell(row=r, column=col, value=float(w.get('volume') or 0)).alignment = _CENTER
        col += 1
        if include_financial:
            volume = float(w.get('volume') or 0)
            salary = float(w.get('current_salary') or 0)
            price = float(w.get('current_price') or 0)
            for value in (salary, volume * salary, price, volume * price):
                ws_works.cell(row=r, column=col, value=round(value, 2)).alignment = _CENTER
                col += 1
        ws_works.cell(
            row=r, column=col,
            value=_author_label(w.get('filled_by_fio'), w.get('filled_by_role')),
        ).alignment = _LEFT
        r += 1
    if r == 2:
        ws_works.cell(row=2, column=1, value='Работы не заполнены').alignment = _LEFT
    _autosize(ws_works)

    # ─────────────────────── Sheet 3: Доп. работы (если есть) ───────────────────────
    extras = sorted(
        report.get('extra_works', []),
        key=lambda row: (str(row.get('object_name') or ''), str(row.get('name') or '')),
    )

    if extras:
        extras_has_teams = any(e.get('team_id') for e in extras)
        ws_extra = wb.create_sheet("Доп. работы")
        headers = ["Объект", "Заявка"] if include_object_context else []
        if extras_has_teams:
            headers.extend(["Бригада", "Наименование", "Ед.изм", "Объём"])
        else:
            headers.extend(["Наименование", "Ед.изм", "Объём"])
        if include_financial:
            headers.extend(["ЗП за ед.", "Сумма ЗП", "Цена за ед.", "Сумма цены"])
        headers.append("Заполнил")
        _write_header(ws_extra, headers)
        r = 2
        for e in extras:
            col = 1
            if include_object_context:
                object_name, application_label = _row_context(e, app_meta)
                ws_extra.cell(row=r, column=col, value=object_name).alignment = _LEFT
                ws_extra.cell(row=r, column=col + 1, value=application_label).alignment = _LEFT
                col += 2
            if extras_has_teams:
                ws_extra.cell(row=r, column=col, value=e.get('team_name') or '—').alignment = _LEFT
                col += 1
            ws_extra.cell(row=r, column=col, value=e.get('name') or '').alignment = _LEFT
            col += 1
            ws_extra.cell(row=r, column=col, value=e.get('unit') or '').alignment = _CENTER
            col += 1
            ws_extra.cell(row=r, column=col, value=float(e.get('volume') or 0)).alignment = _CENTER
            col += 1
            if include_financial:
                volume = float(e.get('volume') or 0)
                salary = float(e.get('salary') or 0)
                price = float(e.get('price') or 0)
                for value in (salary, volume * salary, price, volume * price):
                    ws_extra.cell(row=r, column=col, value=round(value, 2)).alignment = _CENTER
                    col += 1
            ws_extra.cell(
                row=r, column=col,
                value=_author_label(e.get('filled_by_fio'), e.get('filled_by_role')),
            ).alignment = _LEFT
            r += 1
        _autosize(ws_extra)

    if include_financial or include_participant_salary:
        ws_summary = wb.create_sheet("Итоги", 0)
        _write_header(ws_summary, ["Показатель", "Значение"])
        totals = report.get('totals', {})
        summary_rows = [("Всего часов", totals.get('hours', 0))]
        if include_participant_salary:
            summary_rows.append(("ЗП участникам", totals.get('participant_salary', 0)))
        if include_financial:
            summary_rows.extend([
                ("Сумма ЗП по расценкам работ", totals.get('salary', 0)),
                ("Сумма цены", totals.get('price', 0)),
            ])
        for row_no, (label, value) in enumerate(summary_rows, start=2):
            ws_summary.cell(row=row_no, column=1, value=label).alignment = _LEFT
            ws_summary.cell(row=row_no, column=2, value=value).alignment = _CENTER
        _autosize(ws_summary)

    # ─────────────────────── Persist ───────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    blob = buf.read()

    file_objects = list(dict.fromkeys(
        str(context.get('object_name') or '').strip() for context in applications
        if str(context.get('object_name') or '').strip()
    ))
    filename = _build_report_filename(
        object_name=' + '.join(file_objects) if len(file_objects) > 1 else (
            (file_objects[0] if file_objects else None)
            or app_meta.get('object_name') or app_meta.get('object_address')
        ),
        app_id=app_id,
        public_number=app_meta.get('public_number'),
        date_target=app_meta.get('date_target'),
        team_name=team_name,
    )

    return blob, filename


def get_smr_report_targets(report: dict) -> list[tuple[int, str]]:
    """List concrete brigade workbook targets contained in a logical SMR.

    Rows without a brigade belong to the general workbook. They no longer
    create a misleading standalone ``Общие работы`` brigade file.
    """
    rows = [
        *report.get('hours', []),
        *report.get('plan_works', []),
        *report.get('extra_works', []),
    ]
    teams: dict[int, str] = {}
    for row in rows:
        try:
            row_team_id = int(row.get('team_id') or 0)
        except (TypeError, ValueError):
            row_team_id = 0
        if row_team_id > 0:
            teams.setdefault(
                row_team_id,
                str(row.get('team_name') or f'Бригада {row_team_id}'),
            )
    return sorted(teams.items(), key=lambda item: (item[1].casefold(), item[0]))


async def generate_smr_report_files(
    db,
    app_id: int,
    *,
    include_financial: bool = False,
    include_participant_salary: bool = False,
) -> list[tuple[bytes, str]]:
    """Return the general workbook followed by one workbook per brigade.

    The general workbook preserves the original all-in-one report. A merged
    SMR remains convenient: one brigade also receives one workbook even when
    it worked on several source objects. When there is exactly one brigade,
    legacy rows without a brigade tag can be attributed safely and are also
    included in that brigade file. With several brigades such rows stay only
    in the general workbook instead of being guessed or duplicated.
    """
    from smr_data import get_smr_read_model

    report = await get_smr_read_model(db, app_id)
    targets = get_smr_report_targets(report)

    files = [await generate_smr_excel_bytes(
        db,
        app_id,
        include_financial=include_financial,
        include_participant_salary=include_participant_salary,
        team_name='Общий отчёт',
    )]
    include_unassigned = len(targets) == 1
    for target_team_id, target_team_name in targets:
        files.append(await generate_smr_excel_bytes(
            db,
            app_id,
            include_financial=include_financial,
            include_participant_salary=include_participant_salary,
            team_id=target_team_id,
            team_name=target_team_name,
            include_unassigned_for_team=include_unassigned,
        ))
    return files


async def generate_smr_excel_to_disk(db, app_id: int, dest_dir: Path | None = None) -> Path:
    """Variant that persists to `data/uploads/reports` — used when a caller
    wants a sharable URL path instead of a streaming response."""
    blob, filename = await generate_smr_excel_bytes(db, app_id)
    base = dest_dir or Path("data/uploads/reports")
    base.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = base / f"smr_{app_id}_{stamp}.xlsx"
    path.write_bytes(blob)
    return path

import sqlite3
import unittest
import importlib.util
import sys
from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook

from database.hours_repo import HoursRepoMixin
from database.migrations.m_2026_08_smr_participant_salary import run as run_migration
from smr_calculations import SmrNumberError

_WEB_DIR = Path(__file__).resolve().parents[1] / "web"
if str(_WEB_DIR) not in sys.path:
    sys.path.insert(0, str(_WEB_DIR))
_REPORT_SPEC = importlib.util.spec_from_file_location(
    "smr_report_test_module",
    _WEB_DIR / "services" / "smr_report.py",
)
_REPORT_MODULE = importlib.util.module_from_spec(_REPORT_SPEC)
_REPORT_SPEC.loader.exec_module(_REPORT_MODULE)
generate_smr_excel_bytes = _REPORT_MODULE.generate_smr_excel_bytes


class AsyncCursor:
    def __init__(self, cursor):
        self.cursor = cursor

    async def __aenter__(self):
        return self

    async def __aexit__(self, *_args):
        self.cursor.close()

    async def fetchall(self):
        return self.cursor.fetchall()

    async def fetchone(self):
        return self.cursor.fetchone()

    def __await__(self):
        async def _return_self():
            return self
        return _return_self().__await__()


class AsyncConnection:
    def __init__(self):
        self.raw = sqlite3.connect(":memory:")
        self.raw.row_factory = sqlite3.Row

    def execute(self, sql, params=()):
        return AsyncCursor(self.raw.execute(sql, params))

    async def commit(self):
        self.raw.commit()


class Repo(HoursRepoMixin):
    def __init__(self):
        self.conn = AsyncConnection()
        self.conn.raw.executescript(
            """
            CREATE TABLE application_hours (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                app_id INTEGER NOT NULL,
                team_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                hours REAL DEFAULT 0,
                participant_salary REAL DEFAULT 0,
                filled_by_user_id INTEGER,
                filled_at TEXT,
                is_additional INTEGER DEFAULT 0
            );
            CREATE UNIQUE INDEX idx_app_hours_unique
            ON application_hours(app_id, team_id, user_id)
            WHERE is_additional = 0;
            """
        )


class ParticipantSalaryRepoTests(unittest.IsolatedAsyncioTestCase):
    async def test_migration_adds_column_to_existing_hours_table(self):
        conn = AsyncConnection()
        conn.raw.execute(
            "CREATE TABLE application_hours (id INTEGER PRIMARY KEY, hours REAL DEFAULT 0)"
        )
        await run_migration(conn)
        await run_migration(conn)
        columns = {
            row[1] for row in conn.raw.execute("PRAGMA table_info(application_hours)")
        }
        self.assertIn("participant_salary", columns)

    async def test_authorized_save_and_unauthorized_preservation(self):
        repo = Repo()
        item = {"team_id": 2, "user_id": 3, "hours": 8, "participant_salary": "4500,50"}
        await repo.save_app_hours(1, [item], 10, allow_participant_salary=True)

        row = repo.conn.raw.execute(
            "SELECT hours, participant_salary FROM application_hours"
        ).fetchone()
        self.assertEqual(row["hours"], 8)
        self.assertEqual(row["participant_salary"], 4500.5)

        await repo.save_app_hours(
            1,
            [{"team_id": 2, "user_id": 3, "hours": 6, "participant_salary": 1}],
            11,
            allow_participant_salary=False,
        )
        row = repo.conn.raw.execute(
            "SELECT hours, participant_salary FROM application_hours"
        ).fetchone()
        self.assertEqual(row["hours"], 6)
        self.assertEqual(row["participant_salary"], 4500.5)

    async def test_negative_participant_salary_is_rejected(self):
        repo = Repo()
        with self.assertRaisesRegex(SmrNumberError, "ЗП участника"):
            await repo.save_app_hours(
                1,
                [{"team_id": 2, "user_id": 3, "hours": 8, "participant_salary": -1}],
                10,
                allow_participant_salary=True,
            )

    async def test_excel_keeps_participant_salary_separate_from_work_rates(self):
        conn = AsyncConnection()
        conn.raw.executescript(
            """
            CREATE TABLE applications (
                id INTEGER PRIMARY KEY, public_number TEXT, date_target TEXT,
                foreman_name TEXT, object_id INTEGER
            );
            CREATE TABLE objects (id INTEGER PRIMARY KEY, name TEXT, address TEXT);
            INSERT INTO objects VALUES (1, 'Объект', 'Адрес');
            INSERT INTO applications VALUES (1, 'З-100826-01', '2026-08-10', 'Прораб', 1);
            """
        )
        db = type("Db", (), {"conn": conn})()
        report = {
            "hours": [{
                "team_name": "Бригада 1", "fio": "Иванов Иван",
                "specialty": "Монтажник", "hours": 8,
                "participant_salary": 4500.5,
                "filled_by_fio": "Прораб", "filled_by_role": "foreman",
            }],
            "plan_works": [],
            "extra_works": [],
            "totals": {
                "hours": 8, "participant_salary": 4500.5,
                "salary": 9000, "price": 36000,
            },
        }
        with patch("smr_data.get_smr_read_model", AsyncMock(return_value=report)):
            blob, _filename = await generate_smr_excel_bytes(
                db, 1, include_financial=False, include_participant_salary=True
            )

        workbook = load_workbook(BytesIO(blob), data_only=True)
        hours_sheet = workbook["Часы"]
        self.assertEqual(hours_sheet.cell(1, 5).value, "ЗП участника")
        self.assertEqual(hours_sheet.cell(2, 5).value, 4500.5)
        self.assertEqual(workbook["Работы"].max_column, 4)
        summary = workbook["Итоги"]
        self.assertEqual(summary.cell(3, 1).value, "ЗП участникам")
        self.assertEqual(summary.cell(3, 2).value, 4500.5)
        labels = {summary.cell(row, 1).value for row in range(2, summary.max_row + 1)}
        self.assertNotIn("Сумма ЗП по расценкам работ", labels)


if __name__ == "__main__":
    unittest.main()

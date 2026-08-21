import importlib.util
import sqlite3
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from smr_data import get_smr_read_model


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "web") not in sys.path:
    sys.path.insert(0, str(ROOT / "web"))
spec = importlib.util.spec_from_file_location(
    "smr_report_merged_objects", ROOT / "web" / "services" / "smr_report.py"
)
smr_report = importlib.util.module_from_spec(spec)
spec.loader.exec_module(smr_report)


class _Cursor:
    def __init__(self, cursor):
        self.cursor = cursor

    async def __aenter__(self):
        return self

    async def __aexit__(self, *_args):
        self.cursor.close()

    async def fetchone(self):
        return self.cursor.fetchone()

    async def fetchall(self):
        return self.cursor.fetchall()


class _Connection:
    def __init__(self):
        self.raw = sqlite3.connect(":memory:")
        self.raw.row_factory = sqlite3.Row

    def execute(self, sql, params=()):
        return _Cursor(self.raw.execute(sql, params))


def _database():
    conn = _Connection()
    conn.raw.executescript(
        """
        CREATE TABLE applications (
            id INTEGER PRIMARY KEY, public_number TEXT, date_target TEXT,
            foreman_name TEXT, object_id INTEGER, object_address TEXT,
            team_id TEXT, selected_members TEXT, smr_group_id TEXT
        );
        CREATE TABLE objects (id INTEGER PRIMARY KEY, name TEXT, address TEXT);
        CREATE TABLE object_kp_plan (object_id INTEGER, kp_id INTEGER);
        CREATE TABLE kp_catalog (id INTEGER PRIMARY KEY, category TEXT, name TEXT, unit TEXT);
        CREATE TABLE teams (id INTEGER PRIMARY KEY, name TEXT);
        CREATE TABLE users (user_id INTEGER PRIMARY KEY, fio TEXT, role TEXT);
        CREATE TABLE team_members (id INTEGER PRIMARY KEY, fio TEXT, position TEXT);
        CREATE TABLE extra_works_catalog (id INTEGER PRIMARY KEY, name TEXT, unit TEXT);
        CREATE TABLE application_kp (
            id INTEGER PRIMARY KEY, application_id INTEGER, kp_id INTEGER,
            volume REAL, unit TEXT, current_salary REAL, current_price REAL,
            team_id INTEGER, is_additional INTEGER, filled_at TEXT,
            filled_by_user_id INTEGER
        );
        CREATE TABLE application_extra_works (
            id INTEGER PRIMARY KEY, application_id INTEGER, kp_id INTEGER,
            extra_work_id INTEGER, custom_name TEXT, unit TEXT, volume REAL,
            salary REAL, price REAL, team_id INTEGER, is_additional INTEGER,
            filled_at TEXT, filled_by_user_id INTEGER
        );
        CREATE TABLE application_hours (
            id INTEGER PRIMARY KEY, app_id INTEGER, team_id INTEGER,
            user_id INTEGER, hours REAL, participant_salary REAL,
            is_additional INTEGER, filled_at TEXT, filled_by_user_id INTEGER
        );

        INSERT INTO objects VALUES
            (30, 'Ливневая канализация', 'Советской Армии, 83а'),
            (34, 'Водопровод', 'Советской Армии, 83а');
        INSERT INTO applications VALUES
            (235, 'З-100826-01', '2026-08-10', 'Прораб', 30, '', '5', '17,18', 'g'),
            (236, 'З-100826-02', '2026-08-10', 'Прораб', 34, '', '5', '19,20', 'g');
        INSERT INTO kp_catalog VALUES
            (145943, 'Прочее', 'Шурфление', 'м3'),
            (146084, 'Трубы', 'Протаскивание трубы', 'м');
        INSERT INTO object_kp_plan VALUES (30, 145943), (34, 146084);
        INSERT INTO teams VALUES (5, 'Бригада 5');
        INSERT INTO users VALUES (100, 'Прораб', 'foreman');
        INSERT INTO team_members VALUES
            (17, 'Рабочий первого объекта', 'Монтажник'),
            (19, 'Рабочий второго объекта', 'Монтажник');
        -- Consolidated merged storage: every row physically belongs to 235.
        INSERT INTO application_kp VALUES
            (1, 235, 145943, 1.5, 'м3', 10, 20, 5, 0, '2026-08-10', 100),
            (2, 235, 146084, 12.4, 'м', 30, 40, 5, 0, '2026-08-10', 100);
        INSERT INTO application_hours VALUES
            (1, 235, 5, 17, 8, 0, 0, '2026-08-10', 100),
            (2, 235, 5, 19, 7, 0, 0, '2026-08-10', 100);
        """
    )
    return type("Db", (), {"conn": conn})()


def test_read_model_attributes_people_and_works_to_their_objects():
    import asyncio

    async def scenario():
        report = await get_smr_read_model(_database(), 235)
        works = {row['kp_id']: row for row in report['plan_works']}
        hours = {row['member_id']: row for row in report['hours']}
        assert works[145943]['object_name'] == 'Ливневая канализация'
        assert works[146084]['object_name'] == 'Водопровод'
        assert hours[17]['object_name'] == 'Ливневая канализация'
        assert hours[19]['object_name'] == 'Водопровод'

    asyncio.run(scenario())


def test_excel_lists_both_objects_and_separates_rows():
    import asyncio

    async def scenario():
        blob, filename = await smr_report.generate_smr_excel_bytes(_database(), 235)
        workbook = load_workbook(BytesIO(blob), data_only=True)
        assert "Объекты" in workbook.sheetnames
        assert filename.startswith("Ливневая канализация + Водопровод")
        object_rows = {
            workbook['Объекты'].cell(row, 2).value
            for row in range(2, workbook['Объекты'].max_row + 1)
        }
        assert object_rows == {'Ливневая канализация', 'Водопровод'}
        work_rows = {
            workbook['Работы'].cell(row, 4).value: workbook['Работы'].cell(row, 1).value
            for row in range(2, workbook['Работы'].max_row + 1)
        }
        assert work_rows == {
            'Шурфление': 'Ливневая канализация',
            'Протаскивание трубы': 'Водопровод',
        }
        hour_rows = {
            workbook['Часы'].cell(row, 4).value: workbook['Часы'].cell(row, 1).value
            for row in range(2, workbook['Часы'].max_row + 1)
        }
        assert hour_rows == {
            'Рабочий первого объекта': 'Ливневая канализация',
            'Рабочий второго объекта': 'Водопровод',
        }

    asyncio.run(scenario())


def test_same_person_and_work_on_two_objects_are_not_joined_with_slash():
    import asyncio

    async def scenario():
        database = _database()
        database.conn.raw.executescript(
            """
            UPDATE applications SET selected_members='17,21' WHERE id=235;
            UPDATE applications SET selected_members='19,21' WHERE id=236;
            INSERT INTO team_members VALUES (21, 'Общий сотрудник', 'Монтажник');
            INSERT INTO kp_catalog VALUES (150000, 'Общее', 'Одинаковая работа', 'м');
            INSERT INTO object_kp_plan VALUES (30, 150000), (34, 150000);
            INSERT INTO application_kp VALUES
                (3, 235, 150000, 2, 'м', 10, 20, 5, 0, '2026-08-10', 100),
                (4, 236, 150000, 3, 'м', 10, 20, 5, 0, '2026-08-10', 100);
            INSERT INTO application_hours VALUES
                (3, 235, 5, 21, 4, 0, 0, '2026-08-10', 100),
                (4, 236, 5, 21, 5, 0, 0, '2026-08-10', 100);
            """
        )
        report = await get_smr_read_model(database, 235)
        shared_works = [row for row in report['plan_works'] if row['kp_id'] == 150000]
        shared_hours = [row for row in report['hours'] if row['member_id'] == 21]
        assert [row['application_id'] for row in shared_works] == [235, 236]
        assert [row['object_name'] for row in shared_works] == [
            'Ливневая канализация', 'Водопровод',
        ]
        assert [row['hours'] for row in shared_hours] == [4, 5]
        assert [row['object_name'] for row in shared_hours] == [
            'Ливневая канализация', 'Водопровод',
        ]
        assert all('/' not in row['object_name'] for row in shared_works + shared_hours)
        assert report['totals'] == {
            'hours': 24.0,
            'salary': 437.0,
            'participant_salary': 0.0,
            'price': 626.0,
        }

    asyncio.run(scenario())


def test_report_bundle_has_one_workbook_per_brigade_for_merged_smr():
    import asyncio

    async def scenario():
        database = _database()
        database.conn.raw.executescript(
            """
            INSERT INTO teams VALUES (6, 'Бригада 6');
            UPDATE applications SET team_id='6' WHERE id=236;
            UPDATE application_kp SET team_id=6 WHERE id=2;
            UPDATE application_hours SET team_id=6 WHERE id=2;
            """
        )
        files = await smr_report.generate_smr_report_files(database, 235)
        assert len(files) == 3
        report = await get_smr_read_model(database, 235)
        assert smr_report.get_smr_report_targets(report) == [
            (5, 'Бригада 5'),
            (6, 'Бригада 6'),
        ]

        by_name = {filename: load_workbook(BytesIO(blob), data_only=True) for blob, filename in files}
        general_name = next(name for name in by_name if 'Общий отчёт' in name)
        brigade_5_name = next(name for name in by_name if 'Бригада 5' in name)
        brigade_6_name = next(name for name in by_name if 'Бригада 6' in name)
        brigade_5 = by_name[brigade_5_name]
        brigade_6 = by_name[brigade_6_name]
        general = by_name[general_name]

        def value_under(workbook, sheet_name, heading):
            sheet = workbook[sheet_name]
            column = next(
                index for index in range(1, sheet.max_column + 1)
                if sheet.cell(1, index).value == heading
            )
            return sheet.cell(2, column).value

        assert value_under(brigade_5, 'Часы', 'ФИО') == 'Рабочий первого объекта'
        assert value_under(brigade_6, 'Часы', 'ФИО') == 'Рабочий второго объекта'
        assert value_under(brigade_5, 'Работы', 'Наименование') == 'Шурфление'
        assert value_under(brigade_6, 'Работы', 'Наименование') == 'Протаскивание трубы'
        assert brigade_5['Часы'].max_row == 2
        assert brigade_6['Часы'].max_row == 2
        assert general['Часы'].max_row == 3
        assert general['Работы'].max_row == 3

    asyncio.run(scenario())


def test_single_brigade_file_includes_legacy_unassigned_extra_work():
    import asyncio

    async def scenario():
        database = _database()
        database.conn.raw.executescript(
            """
            INSERT INTO extra_works_catalog VALUES (9, 'Газорезка', 'ч');
            INSERT INTO application_extra_works
                (id, application_id, kp_id, extra_work_id, custom_name, unit,
                 volume, salary, price, team_id, is_additional, filled_at,
                 filled_by_user_id)
            VALUES (1, 235, NULL, 9, 'Газорезка', 'ч', 2, 100, 200,
                    NULL, 0, '2026-08-10', 100);
            """
        )
        files = await smr_report.generate_smr_report_files(database, 235)
        assert len(files) == 2
        by_name = {
            filename: load_workbook(BytesIO(blob), data_only=True)
            for blob, filename in files
        }
        brigade_name = next(name for name in by_name if 'Бригада 5' in name)
        brigade = by_name[brigade_name]
        extra_sheet = brigade['Доп. работы']
        name_column = next(
            index for index in range(1, extra_sheet.max_column + 1)
            if extra_sheet.cell(1, index).value == 'Наименование'
        )
        team_column = next(
            index for index in range(1, extra_sheet.max_column + 1)
            if extra_sheet.cell(1, index).value == 'Бригада'
        )
        assert extra_sheet.cell(2, name_column).value == 'Газорезка'
        assert extra_sheet.cell(2, team_column).value == 'Бригада 5'

    asyncio.run(scenario())


def test_unassigned_extra_is_not_guessed_between_multiple_brigades():
    import asyncio

    async def scenario():
        database = _database()
        database.conn.raw.executescript(
            """
            INSERT INTO teams VALUES (6, 'Бригада 6');
            UPDATE applications SET team_id='6' WHERE id=236;
            UPDATE application_kp SET team_id=6 WHERE id=2;
            UPDATE application_hours SET team_id=6 WHERE id=2;
            INSERT INTO extra_works_catalog VALUES (9, 'Газорезка', 'ч');
            INSERT INTO application_extra_works
                (id, application_id, kp_id, extra_work_id, custom_name, unit,
                 volume, salary, price, team_id, is_additional, filled_at,
                 filled_by_user_id)
            VALUES (1, 235, NULL, 9, 'Газорезка', 'ч', 2, 100, 200,
                    NULL, 0, '2026-08-10', 100);
            """
        )
        files = await smr_report.generate_smr_report_files(database, 235)
        assert len(files) == 3
        by_name = {
            filename: load_workbook(BytesIO(blob), data_only=True)
            for blob, filename in files
        }
        general_name = next(name for name in by_name if 'Общий отчёт' in name)
        assert 'Доп. работы' in by_name[general_name].sheetnames
        for filename, workbook in by_name.items():
            if filename == general_name:
                continue
            assert 'Доп. работы' not in workbook.sheetnames

    asyncio.run(scenario())
